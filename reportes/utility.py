import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.compat import range
from openpyxl.styles import Font, Fill, Alignment, Border, PatternFill, Side, GradientFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, Rule
from sga.settings import BASE_DIR

class GenerateXlsxReport():
    """docstring for GenerateXlsReport"""
    def __init__(self,xlsx_filename, columns, reportData,sheet_name):
        dir_name = os.path.join(BASE_DIR, "staticfiles/custom/")
        files = os.listdir(dir_name)

        for item in files:
            if item.endswith(".xlsx"):
                os.remove(os.path.join(dir_name, item))

        self.xlsx_filename = os.path.join(dir_name,xlsx_filename)
        self.queryset_data = reportData
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Report"
        cell = self.ws.cell(row=1, column=1)
        cell.value = sheet_name 
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(b=True, color="FFFFFF")
        cell.fill = GradientFill(stop=("000000", "FFFFFF"))
        double = Side(border_style="double", color="FFFFFF")
        thin = Side(border_style="thin", color="000000")
        self.border = Border(top=double, left=thin, right=thin, bottom=double)
        cell.border = self.border
        self.columns = columns
        self.ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=len(self.columns))

    def generate(self):
        row_num =2
        for col_num in range(len(self.columns)):
            c = self.ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = self.columns[col_num][0]
            c.font= Font(size=15, bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            self.ws.column_dimensions[get_column_letter(col_num+1)].width = self.columns[col_num][1]
            self.ws.row_dimensions[row_num+1].height = 40
        for row in self.queryset_data:
            row_num += 1
            for col_num in range(len(row)):
                c = self.ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = row[col_num]
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            self.ws.row_dimensions[row_num+1].height = 20

        
        print(self.xlsx_filename)
        self.wb.save("{}".format(self.xlsx_filename))
        return self.xlsx_filename
